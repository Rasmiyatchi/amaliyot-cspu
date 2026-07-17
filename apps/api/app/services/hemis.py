"""Talabalar Excel import service.

Namuna shablon ustunlari (13 ta): To'liq ismi, Viloyat, Tuman, Jins, Kurs,
Fakultet, Guruh, Ta'lim tili, O'quv yili, Semestr, Bitiruvchi, Mutaxassislik,
Ta'lim shakli. "Talaba/Amaliyot id" ustuni IXTIYORIY — berilmasa tizim
avtomatik generatsiya qiladi. Majburiy: To'liq ismi, Mutaxassislik, Guruh, Kurs.
"""

import io
import re
import secrets
import string
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.academic import AcademicYear, Direction, Group
from app.models.enums import DegreeType, EducationForm, Gender, StudentStatus, UserRole
from app.models.student import Student
from app.models.user import User
from app.schemas.hemis import HemisCredentials, HemisImportError, HemisImportResponse

# Header normalizatsiya: Unicode apostrofni ASCII'ga
_UNICODE_APOSTROPHES = {ord("‘"): "'", ord("’"): "'", ord("ʻ"): "'", ord("ʼ"): "'"}


def _normalize_header(s: str) -> str:
    return s.translate(_UNICODE_APOSTROPHES).strip().lower()


# Header (normalized) → canonical key
HEADER_MAP: dict[str, str] = {
    # "talaba id" / "hemis id" — IXTIYORIY (legacy). Berilsa amaliyot id sifatida
    # ishlatiladi, berilmasa tizim avtomatik generatsiya qiladi.
    "talaba id": "amaliyot_id",
    "hemis id": "amaliyot_id",
    "amaliyot id": "amaliyot_id",
    "to'liq ismi": "full_name",
    "viloyat": "region",
    "tuman": "district",
    "jins": "gender",
    # Maxfiylik: tug'ilgan sana / pasport / JSHSHIR ustunlari bo'lsa ham
    # e'tiborsiz qoldiriladi (saqlanmaydi) — foydalanuvchi talabi 2026-06-20
    "kurs": "course",
    "fakultet": "faculty_name",  # info only — Direction orqali topamiz
    "guruh": "group_name",
    "ta'lim tili": "education_language",
    "o'quv yili": "academic_year_name",
    "semestr": "semester",
    "bitiruvchi": "is_graduating",
    "mutaxassislik": "direction_code",
    "ta'lim turi": "degree_type",
    "ta'lim shakli": "education_form",
}

# "amaliyot_id" MAJBURIY EMAS — berilmasa tizim avtomatik generatsiya qiladi.
REQUIRED_KEYS = {"full_name", "direction_code", "group_name", "course"}


async def _generate_unique_login(db: AsyncSession, prefix: str) -> str:
    """{prefix}{8 random digits} formatida unikal login generatsiya qiladi.

    Masalan prefix='2500' → '250012345678'.
    Konflikt bo'lsa qayta urinadi (10 marta), keyin RuntimeError tashlaydi.
    """
    for _ in range(10):
        suffix = "".join(secrets.choice(string.digits) for _ in range(8))
        candidate = f"{prefix}{suffix}"
        existing = (
            await db.execute(select(User.id).where(User.username == candidate))
        ).scalar_one_or_none()
        if not existing:
            return candidate
    raise RuntimeError("Unique login generatsiya qila olmadim — qayta urinib ko'ring")


def _parse_full_name(full: str) -> tuple[str, str, str | None]:
    """'ALIYEVA SHAXZODA AXTAM QIZI' → ('ALIYEVA', 'SHAXZODA', 'AXTAM QIZI')"""
    parts = full.strip().split()
    if len(parts) < 2:
        raise ValueError(f"Ism kamida 2 so'z bo'lishi kerak: {full!r}")
    last_name = parts[0]
    first_name = parts[1]
    middle_name = " ".join(parts[2:]) if len(parts) > 2 else None
    return last_name, first_name, middle_name


def _parse_course(v: Any) -> int:
    if v is None:
        raise ValueError("kurs bo'sh")
    s = str(v).strip()
    m = re.match(r"(\d+)", s)  # "3-kurs" → 3
    if not m:
        raise ValueError(f"kurs noto'g'ri: {v!r}")
    c = int(m.group(1))
    if not (1 <= c <= 5):
        raise ValueError(f"kurs 1-5 bo'lishi kerak: {c}")
    return c


def _parse_semester(v: Any) -> int | None:
    if v is None:
        return None
    s = str(v).strip()
    m = re.match(r"(\d+)", s)  # "5-semestr" → 5
    return int(m.group(1)) if m else None


def _parse_gender(v: Any) -> Gender | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"erkak", "er", "male", "m", "o'g'il"}:
        return Gender.MALE
    if s in {"ayol", "female", "f", "qiz"}:
        return Gender.FEMALE
    return None


def _parse_education_form(v: Any) -> EducationForm | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    # Substring tekshiruvi — "Ikkinchi oliy (sirtqi)" kabi qiymatlarni ham ushlaydi.
    # "masofaviy" "sirtqi"dan oldin tekshiriladi (ba'zan birga yoziladi).
    if "kunduzgi" in s:
        return EducationForm.DAYTIME
    if "kechki" in s:
        return EducationForm.EVENING
    if "masofaviy" in s:
        return EducationForm.DISTANCE
    if "sirtqi" in s:
        return EducationForm.CORRESPONDENCE
    return None


def _parse_degree_type(v: Any) -> DegreeType | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if "bakalavr" in s:
        return DegreeType.BACHELOR
    if "magistr" in s:
        return DegreeType.MASTER
    if "phd" in s or "doktor" in s:
        return DegreeType.PHD
    return None


def _parse_bool(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"ha", "yes", "true", "1"}


def _parse_excel(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], list[HemisImportError]]:
    errors: list[HemisImportError] = []
    rows: list[dict[str, Any]] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        errors.append(HemisImportError(row=0, message=f"Excel ochib bo'lmadi: {e}"))
        return rows, errors

    ws = wb.active
    if ws is None:
        errors.append(HemisImportError(row=0, message="Excel bo'sh"))
        return rows, errors

    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        errors.append(HemisImportError(row=1, message="Header yo'q"))
        return rows, errors

    # Headerlarni normalize qilib kanonik kalitga map qilamiz
    col_to_key: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        norm = _normalize_header(str(cell))
        if norm in HEADER_MAP:
            col_to_key[idx] = HEADER_MAP[norm]

    mapped_keys = set(col_to_key.values())
    missing = REQUIRED_KEYS - mapped_keys
    if missing:
        errors.append(
            HemisImportError(
                row=1,
                message=f"Majburiy ustunlar yetmaydi: {', '.join(sorted(missing))}",
            )
        )
        return rows, errors

    for row_idx, row in enumerate(iterator, start=2):
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        rec: dict[str, Any] = {"_row_idx": row_idx}
        for col_idx, key in col_to_key.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                val = val.strip() or None
            rec[key] = val
        rows.append(rec)

    return rows, errors


async def import_students(db: AsyncSession, file_bytes: bytes) -> HemisImportResponse:
    rows, parse_errors = _parse_excel(file_bytes)
    if parse_errors:
        return HemisImportResponse(
            total_rows=0, created=0, skipped=0, errors=parse_errors, credentials=[]
        )

    ay = (
        await db.execute(select(AcademicYear.id).where(AcademicYear.is_active.is_(True)))
    ).scalar_one_or_none()
    if not ay:
        return HemisImportResponse(
            total_rows=len(rows),
            created=0,
            skipped=0,
            errors=[HemisImportError(row=0, message="Aktiv akademik yil topilmadi")],
            credentials=[],
        )
    ay_id = ay  # plain UUID — commit'lardan keyin eskirmaydi

    from app.core.config import settings as app_settings

    # Kesh — ORM obyekt EMAS, ID saqlaymiz (per-row commit'da eskirmaydi).
    direction_cache: dict[str, UUID | None] = {}
    group_cache: dict[tuple[str, str], UUID] = {}

    errors: list[HemisImportError] = []
    credentials: list[HemisCredentials] = []
    created = 0
    skipped = 0

    for rec in rows:
        row_idx = rec["_row_idx"]
        raw_id = rec.get("amaliyot_id")
        incoming_id = str(raw_id).strip() if raw_id else None

        try:
            # ── Amaliyot id bo'yicha dublikat (read-only) ──
            if incoming_id and (
                await db.execute(select(Student.id).where(Student.hemis_id == incoming_id))
            ).first():
                skipped += 1
                continue

            # ── Yo'nalish — Mutaxassislik shifri (kodi) bo'yicha ──
            direction_code = str(rec.get("direction_code", "")).strip()
            if direction_code not in direction_cache:
                # Kod unikal EMAS (bir kodda bir nechta yo'nalish bo'lishi mumkin) —
                # birinchisini olamiz (shablonda faqat kod bo'lgani uchun).
                direction_cache[direction_code] = (
                    await db.execute(
                        select(Direction.id).where(Direction.code == direction_code)
                    )
                ).scalars().first()
            dir_id = direction_cache[direction_code]
            if not dir_id:
                shown = direction_code or "(bo'sh)"
                raise ValueError(
                    f"Yo'nalish topilmadi — bu shifr (kod) akademik tuzilmada yo'q: {shown}"
                )

            course = _parse_course(rec.get("course"))

            group_name = str(rec.get("group_name", "")).strip()
            if not group_name:
                raise ValueError("Guruh nomi bo'sh")

            # ── Guruh — yo'q bo'lsa avto-yaratamiz (ID-cache) ──
            gkey = (str(dir_id), group_name)
            group_id = group_cache.get(gkey)
            if group_id is None:
                g_id = (
                    await db.execute(
                        select(Group.id).where(
                            Group.direction_id == dir_id,
                            Group.academic_year_id == ay_id,
                            Group.name == group_name,
                        )
                    )
                ).scalar_one_or_none()
                if g_id is None:
                    g = Group(
                        direction_id=dir_id,
                        academic_year_id=ay_id,
                        name=group_name,
                        course=course,
                    )
                    db.add(g)
                    await db.flush()
                    g_id = g.id
                group_id = g_id

            # ── Ism ──
            full_name_raw = str(rec.get("full_name", "")).strip()
            if not full_name_raw:
                raise ValueError("To'liq ismi bo'sh")
            last_name, first_name, middle_name = _parse_full_name(full_name_raw)

            # ── Ism+guruh bo'yicha dublikat (amaliyot id yo'q bo'lsa) ──
            if not incoming_id:
                dup = (
                    await db.execute(
                        select(Student.id)
                        .join(User, Student.user_id == User.id)
                        .where(
                            Student.group_id == group_id,
                            func.lower(User.last_name) == last_name.lower(),
                            func.lower(User.first_name) == first_name.lower(),
                        )
                    )
                ).first()
                if dup:
                    skipped += 1
                    continue

            # ── Login + user + student ──
            generated_login = await _generate_unique_login(
                db, app_settings.LOGIN_YEAR_PREFIX
            )
            password = generated_login  # login va parol bir xil
            amaliyot_id = incoming_id or generated_login

            user = User(
                username=generated_login,
                password_hash=hash_password(password),
                role=UserRole.STUDENT,
                is_active=True,
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                must_change_password=True,
            )
            db.add(user)
            await db.flush()

            student = Student(
                user_id=user.id,
                hemis_id=amaliyot_id,
                gender=_parse_gender(rec.get("gender")),
                region=rec.get("region") or None,
                district=rec.get("district") or None,
                group_id=group_id,
                current_semester=_parse_semester(rec.get("semester")),
                is_graduating=_parse_bool(rec.get("is_graduating")),
                education_language=rec.get("education_language") or None,
                education_form=_parse_education_form(rec.get("education_form")),
                degree_type=_parse_degree_type(rec.get("degree_type")),
                status=StudentStatus.STUDYING,
            )
            db.add(student)

            # ── Shu qatorni ATOMAR commit — xato bo'lsa faqat shu qator yo'qoladi ──
            await db.commit()

            group_cache[gkey] = group_id  # muvaffaqiyatdan keyin keshlaymiz
            credentials.append(
                HemisCredentials(
                    amaliyot_id=amaliyot_id,
                    full_name=full_name_raw,
                    group_name=group_name,
                    course=course,
                    direction_code=direction_code,
                    username=generated_login,
                    password=password,
                )
            )
            created += 1

        except Exception as e:  # noqa: BLE001
            # Faqat shu qatorni bekor qilamiz (oldingi qatorlar allaqachon commit qilingan).
            await db.rollback()
            errors.append(
                HemisImportError(row=row_idx, amaliyot_id=incoming_id, message=str(e))
            )
            continue

    return HemisImportResponse(
        total_rows=len(rows),
        created=created,
        skipped=skipped,
        errors=errors,
        credentials=credentials,
    )
