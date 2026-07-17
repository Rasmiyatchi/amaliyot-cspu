"""O'qituvchi (amaliyot rahbari) excel import service.

Namuna shablon ustunlari: FISh (yagona ustun), Fakultet, Kafedra, Lavozim,
E-pochta. Header qatori dinamik aniqlanadi (yuqorida sarlavha/izoh qatorlari
bo'lsa ham). FISh yagona ustun bo'lsa Familiya/Ism/Sharifga bo'linadi; eski
"Familiya"+"Ism" ustunlari ham qo'llab-quvvatlanadi.

Login berilmasa avtomatik generatsiya (LOGIN_YEAR_PREFIX), parol = login,
must_change_password=True. Fakultet/kafedra nomi bo'yicha topiladi (kafedra
yo'q bo'lsa avto-yaratiladi). Tashkilotlar nomi bo'yicha topiladi.
"""

import io
import re
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings as app_settings
from app.core.security import hash_password
from app.models.academic import Department, Faculty
from app.models.enums import UserRole
from app.models.organization import Organization
from app.models.supervisor import Supervisor, SupervisorOrganization
from app.models.user import User
from app.schemas.supervisor_import import (
    SupervisorImportCredentials,
    SupervisorImportError,
    SupervisorImportResponse,
)
from app.services.hemis import (
    _generate_unique_login,
    _normalize_header,
    _parse_full_name,
)

HEADER_MAP: dict[str, str] = {
    # Yagona FISh ustuni (namuna shablon) — Familiya/Ism/Sharifga bo'linadi
    "fish (to'liq ismi sharifi)": "full_name",
    "fish": "full_name",
    "f.i.sh.": "full_name",
    "f.i.sh": "full_name",
    "f.i.o": "full_name",
    "to'liq ismi": "full_name",
    "to'liq ismi sharifi": "full_name",
    # Eski alohida ustunlar (orqaga moslik)
    "familiya": "last_name",
    "ism": "first_name",
    "otasining ismi": "middle_name",
    "otasi": "middle_name",
    "lavozim": "position",
    "telefon": "phone",
    "tel": "phone",
    "email": "email",
    "e-pochta": "email",
    "e pochta": "email",
    "mutaxassislik": "specialty",
    "tajriba": "experience_years",
    "fakultet": "faculty_name",
    "kafedra": "department_name",
    "tashkilot": "organization_names",
    "login": "username",
}


def _parse_excel(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], list[SupervisorImportError]]:
    errors: list[SupervisorImportError] = []
    rows: list[dict[str, Any]] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        errors.append(SupervisorImportError(row=0, message=f"Excel ochib bo'lmadi: {e}"))
        return rows, errors

    ws = wb.active
    if ws is None:
        errors.append(SupervisorImportError(row=0, message="Excel bo'sh"))
        return rows, errors

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        errors.append(SupervisorImportError(row=0, message="Excel bo'sh"))
        return rows, errors

    # Header qatorini dinamik aniqlaymiz: yuqorida sarlavha/izoh qatorlari
    # bo'lsa ham — kamida 2 ta ustun nomi tanilgan birinchi qator = header.
    header_idx: int | None = None
    col_to_key: dict[int, str] = {}
    for i, r in enumerate(all_rows[:15]):
        mapping: dict[int, str] = {}
        for idx, cell in enumerate(r):
            if cell is None:
                continue
            norm = _normalize_header(str(cell))
            if norm in HEADER_MAP:
                mapping[idx] = HEADER_MAP[norm]
        if len(mapping) >= 2:
            header_idx = i
            col_to_key = mapping
            break

    if header_idx is None:
        errors.append(
            SupervisorImportError(row=1, message="Ustun nomlari (header) topilmadi")
        )
        return rows, errors

    keys = set(col_to_key.values())
    has_name = "full_name" in keys or {"last_name", "first_name"} <= keys
    if not has_name:
        errors.append(
            SupervisorImportError(
                row=header_idx + 1,
                message="Majburiy ustun yetmaydi: FISh (yoki Familiya + Ism)",
            )
        )
        return rows, errors

    # Header'dan keyingi qatorlar — data (Excel qator raqami = index + 1)
    for j in range(header_idx + 1, len(all_rows)):
        row = all_rows[j]
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        rec: dict[str, Any] = {"_row_idx": j + 1}
        for col_idx, key in col_to_key.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                val = val.strip() or None
            rec[key] = val
        rows.append(rec)
    return rows, errors


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _clean_email(value: Any) -> str | None:
    """Excel'dagi email — faqat to'g'ri formatdagisi saqlanadi.

    Aks holda None (masalan "yo'q", "-", ism). Buzuq qiymat DB'ga tushsa,
    SupervisorRead butun ro'yxatni 500 qilardi.
    """
    if not value:
        return None
    s = str(value).strip()
    return s if _EMAIL_RE.match(s) else None


async def _find_faculty(db: AsyncSession, name: str) -> Faculty | None:
    return (
        await db.execute(
            select(Faculty).where(func.lower(Faculty.name) == name.strip().lower())
        )
    ).scalar_one_or_none()


async def _find_or_create_department(
    db: AsyncSession, faculty_id: Any, name: str
) -> Department:
    dept = (
        await db.execute(
            select(Department).where(
                Department.faculty_id == faculty_id,
                func.lower(Department.name) == name.strip().lower(),
            )
        )
    ).scalar_one_or_none()
    if dept:
        return dept
    dept = Department(faculty_id=faculty_id, name=name.strip())
    db.add(dept)
    await db.flush()
    return dept


async def _resolve_organizations(db: AsyncSession, raw: str) -> list[Any]:
    """Vergul/nuqta-vergul bilan ajratilgan nomlardan organization id'lar (max 5)."""
    names = [n.strip() for n in raw.replace(";", ",").split(",") if n.strip()]
    org_ids: list[Any] = []
    for nm in names[:5]:
        org = (
            await db.execute(
                select(Organization.id).where(
                    func.lower(Organization.name) == nm.lower()
                )
            )
        ).scalar_one_or_none()
        if org:
            org_ids.append(org)
    return org_ids


async def import_supervisors(
    db: AsyncSession, file_bytes: bytes
) -> SupervisorImportResponse:
    rows, parse_errors = _parse_excel(file_bytes)
    errors = list(parse_errors)
    credentials: list[SupervisorImportCredentials] = []
    created = 0
    skipped = 0

    for rec in rows:
        row_idx = rec.get("_row_idx", 0)
        last_name = rec.get("last_name")
        first_name = rec.get("first_name")
        middle_name = rec.get("middle_name")
        full_name_raw = rec.get("full_name")

        # Yagona FISh ustuni bo'lsa — Familiya / Ism / Sharifga bo'lamiz.
        if full_name_raw and not (last_name and first_name):
            try:
                last_name, first_name, middle_name = _parse_full_name(str(full_name_raw))
            except ValueError:
                parts = str(full_name_raw).split()
                last_name = parts[0] if parts else None
                first_name = parts[1] if len(parts) > 1 else None
                middle_name = " ".join(parts[2:]) if len(parts) > 2 else None

        full_name = (
            str(full_name_raw).strip()
            if full_name_raw
            else f"{last_name or ''} {first_name or ''}".strip()
        )

        if not last_name or not first_name:
            errors.append(
                SupervisorImportError(
                    row=row_idx,
                    name=full_name or None,
                    message="FISh (Familiya + Ism) majburiy",
                )
            )
            continue

        # Login band bo'lsa — o'tkazib yuboramiz (savepointdan tashqari, read-only)
        username = rec.get("username")
        if username:
            username = str(username).strip()
            exists = (
                await db.execute(select(User.id).where(User.username == username))
            ).scalar_one_or_none()
            if exists:
                skipped += 1
                continue

        try:
            # Har bir qator alohida SAVEPOINT'da — xato qator boshqalarni buzmaydi
            async with db.begin_nested():
                faculty_id = None
                department_id = None
                faculty_name = rec.get("faculty_name")
                if faculty_name:
                    faculty = await _find_faculty(db, str(faculty_name))
                    if not faculty:
                        raise ValueError(f"Fakultet topilmadi: {faculty_name}")
                    faculty_id = faculty.id
                    dept_name = rec.get("department_name")
                    if dept_name:
                        dept = await _find_or_create_department(
                            db, faculty_id, str(dept_name)
                        )
                        department_id = dept.id

                login = username or await _generate_unique_login(
                    db, app_settings.LOGIN_YEAR_PREFIX
                )
                password = login  # login = parol (birinchi kirishda almashtiriladi)

                user = User(
                    username=login,
                    password_hash=hash_password(password),
                    role=UserRole.SUPERVISOR,
                    is_active=True,
                    first_name=str(first_name),
                    last_name=str(last_name),
                    middle_name=(str(middle_name) if middle_name else None),
                    phone=(str(rec["phone"]) if rec.get("phone") else None),
                    email=_clean_email(rec.get("email")),
                    must_change_password=True,
                )
                db.add(user)
                await db.flush()

                exp = rec.get("experience_years")
                try:
                    exp_val = int(exp) if exp is not None else None
                except (ValueError, TypeError):
                    exp_val = None

                supervisor = Supervisor(
                    user_id=user.id,
                    position=str(rec.get("position") or "O'qituvchi"),
                    specialty=(str(rec["specialty"]) if rec.get("specialty") else None),
                    experience_years=exp_val,
                    faculty_id=faculty_id,
                    department_id=department_id,
                    capacity=5,
                )
                db.add(supervisor)
                await db.flush()

                org_raw = rec.get("organization_names")
                if org_raw:
                    for org_id in await _resolve_organizations(db, str(org_raw)):
                        db.add(
                            SupervisorOrganization(
                                supervisor_id=supervisor.id, organization_id=org_id
                            )
                        )

            credentials.append(
                SupervisorImportCredentials(
                    full_name=full_name, username=login, password=password
                )
            )
            created += 1
        except Exception as e:  # noqa: BLE001
            logger.warning(f"Supervisor import xato (qator {row_idx}): {e}")
            errors.append(
                SupervisorImportError(
                    row=row_idx, name=full_name, message=f"Xatolik: {e}"
                )
            )

    await db.commit()
    return SupervisorImportResponse(
        total_rows=len(rows),
        created=created,
        skipped=skipped,
        errors=errors,
        credentials=credentials,
    )
