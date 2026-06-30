"""Import uchun namuna Excel shablonlarini generatsiya qiladi.

Ustun ta'riflari parserlar bilan (hemis.py, supervisor_import.py) mos.
Namuna (sariq) qator foydalanuvchi tomonidan o'chirilishi kerak.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")

# ─── Talaba shabloni (13 ustun) ───────────────────────────
_STUDENT_HEADERS = [
    "To'liq ismi",
    "Viloyat",
    "Tuman",
    "Jins",
    "Kurs",
    "Fakultet",
    "Guruh",
    "Ta'lim tili",
    "O'quv yili",
    "Semestr",
    "Bitiruvchi",
    "Mutaxassislik",
    "Ta'lim shakli",
]
_STUDENT_SAMPLE = [
    "ALIYEVA SHAXZODA AXTAM QIZI",
    "Toshkent viloyati",
    "Parkent tumani",
    "Ayol",
    "4-kurs",
    "Gumanitar fanlar",
    "GF-21",
    "O'zbek",
    "2025-2026",
    "8-semestr",
    "Yo'q",
    "5111200",
    "Kunduzgi",
]

# ─── O'qituvchi shabloni (5 ustun) ────────────────────────
_SUPERVISOR_TITLE = "O'QITUVCHILARNI IMPORT QILISH SHABLONI"
_SUPERVISOR_NOTE = (
    "Sariq (namuna) qatorni o'chiring va o'z ma'lumotlaringizni kiriting. "
    "FISh, Fakultet, Mutaxassislik majburiy."
)
_SUPERVISOR_HEADERS = ["FISh (To'liq ismi sharifi)", "Fakultet", "Kafedra", "Lavozim", "E-pochta"]
_SUPERVISOR_SAMPLE = [
    "Karimov Sardor Aliyevich",
    "Aniq fanlar fakulteti",
    "Algebra va matematik analiz kafedrasi",
    "Dotsent",
    "s.karimov@cspu.uz",
]


def _autosize(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, min(40, len(h) + 6))


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_students_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    for i, h in enumerate(_STUDENT_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, v in enumerate(_STUDENT_SAMPLE, start=1):
        ws.cell(row=2, column=i, value=v).fill = _SAMPLE_FILL
    _autosize(ws, _STUDENT_HEADERS)
    ws.freeze_panes = "A2"
    return _to_bytes(wb)


_CREDENTIALS_HEADERS = [
    "Amaliyot ID",
    "F.I.SH.",
    "Guruh",
    "Kurs",
    "Mutaxassislik",
    "Login",
    "Bir martalik parol",
]


def build_credentials_xlsx(credentials: list) -> bytes:
    """Import javobidagi login/parollarni to'liq ma'lumot bilan Excel'ga yozadi.

    `credentials` — HemisCredentials obyektlari ro'yxati (yoki shu maydonli dict'lar).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Login va parollar"
    for i, h in enumerate(_CREDENTIALS_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _get(obj, key):
        return obj.get(key) if isinstance(obj, dict) else getattr(obj, key, None)

    for r, cred in enumerate(credentials, start=2):
        ws.cell(row=r, column=1, value=_get(cred, "amaliyot_id"))
        ws.cell(row=r, column=2, value=_get(cred, "full_name"))
        ws.cell(row=r, column=3, value=_get(cred, "group_name"))
        ws.cell(row=r, column=4, value=_get(cred, "course"))
        ws.cell(row=r, column=5, value=_get(cred, "direction_code"))
        ws.cell(row=r, column=6, value=_get(cred, "username"))
        ws.cell(row=r, column=7, value=_get(cred, "password"))

    _autosize(ws, _CREDENTIALS_HEADERS)
    ws.freeze_panes = "A2"
    return _to_bytes(wb)


_RECORDS_HEADERS = [
    "№",
    "Talaba",
    "Mutaxassislik",
    "Guruh",
    "Kurs",
    "Amaliyot nomi",
    "Korxona",
    "Muddat",
    "Davomat %",
    "Korxona bahosi",
    "Qaydnoma bahosi",
]


def build_records_xlsx(rows: list) -> bytes:
    """Qaydnomalar ro'yxatini Excel'ga yozadi."""

    def _g(obj, key):
        return obj.get(key) if isinstance(obj, dict) else getattr(obj, key, None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Qaydnomalar"
    for i, h in enumerate(_RECORDS_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, r in enumerate(rows, start=1):
        sd, ed = _g(r, "start_date"), _g(r, "end_date")
        muddat = f"{sd} — {ed}" if sd and ed else ""
        att = _g(r, "attendance_pct")
        kg, kgm = _g(r, "korxona_grade"), _g(r, "korxona_grade_max")
        korxona = f"{kg}/{kgm}" if kg is not None and kgm else (kg if kg is not None else "")
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=_g(r, "student_name"))
        ws.cell(row=row, column=3, value=_g(r, "direction_name") or _g(r, "direction_code"))
        ws.cell(row=row, column=4, value=_g(r, "group_name"))
        ws.cell(row=row, column=5, value=_g(r, "course"))
        ws.cell(row=row, column=6, value=_g(r, "practice_type_name"))
        ws.cell(row=row, column=7, value=_g(r, "object_name"))
        ws.cell(row=row, column=8, value=muddat)
        ws.cell(row=row, column=9, value=f"{att}%" if att is not None else "")
        ws.cell(row=row, column=10, value=korxona)
        ws.cell(row=row, column=11, value=_g(r, "qaydnoma_grade"))

    _autosize(ws, _RECORDS_HEADERS)
    ws.freeze_panes = "A2"
    return _to_bytes(wb)


def build_supervisors_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "O'qituvchilar"
    ncols = len(_SUPERVISOR_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1, value=_SUPERVISOR_TITLE)
    title.font = Font(bold=True, size=13)
    title.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    note = ws.cell(row=2, column=1, value=_SUPERVISOR_NOTE)
    note.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, h in enumerate(_SUPERVISOR_HEADERS, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, v in enumerate(_SUPERVISOR_SAMPLE, start=1):
        ws.cell(row=4, column=i, value=v).fill = _SAMPLE_FILL
    _autosize(ws, _SUPERVISOR_HEADERS)
    return _to_bytes(wb)
